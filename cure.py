from re import T
import tkinter as tk
from turtle import bgcolor, width
import docxtpl
import pyautogui
import win32com
import time
import os
import pandas as pd
import datetime
import locale
import subprocess
import win32gui
import openpyxl

from win32.win32gui import *
from datetime import date, timedelta
from datetime import datetime
from tkcalendar import DateEntry
from tkinter import ttk, filedialog
from docxtpl import DocxTemplate
from win32com import client
from decimal import Decimal
from pynput import keyboard
from pynput.keyboard import Controller, Key
from datetime import datetime
from dateutil.relativedelta import relativedelta
from pynput.mouse import Listener
from openpyxl import load_workbook, Workbook


def resetVars():
    try:
        with open("path.txt", "r") as file:
            lines = file.readlines()
            if len(lines) >= 1:
                first_line = lines[0].strip()
                path_var = tk.StringVar()
                path_var.set(first_line)
                
    except FileNotFoundError:
        pass

# Function to handle the "Change" button for initials
def changeInitials():
    new_initials = initials_input.get()
    # Write the new initials to the info.txt file
    with open("info.txt", "w") as file:
        file.write(new_initials)

word = client.Dispatch("Word.Application")

window = tk.Tk()
tab_control = ttk.Notebook(window)
window.geometry("621x590")
window.title("Document Creator")


# Just simply import the azure.tcl file
window.tk.call("source", "azure.tcl")

# Then set the theme you want with the set_theme procedure
window.tk.call("set_theme", "light")


# Create a top frame to go underneath the notebook with a light gray background
topFrame = tk.Frame(window, bg="lightgray", height=15)
topFrame.pack(side="top", fill="both")

# Label 'Initials'
initials_label = tk.Label(topFrame, text="Initials:", fg="black")
initials_label.grid(row=0, column=1, padx=5, pady=5, sticky="w")

# Text input for initials
initials_input = tk.Entry(topFrame, width=5)
initials_input.grid(row=0, column=2, padx=5, pady=5, sticky="w")

# Button to change initials
change_button = tk.Button(topFrame, text="Change", command=changeInitials)
change_button.grid(row=0, column=3, padx=5, pady=5, sticky="w")

# Function to read the initials from info.txt and set it as the default value in the text input
def setVars():
    try:
        with open("info.txt", "r") as file:
            lines = file.readlines()
            if len(lines) >= 1:
                first_line = lines[0].strip()
                initials_input.delete(0, tk.END)
                initials_input.insert(0, first_line)
    except FileNotFoundError:
        pass

setVars()

def create_tab(tab_control, text):
    frame = ttk.Frame(tab_control)
    tab_control.add(frame, text=text)
    return frame

# Create tabs

tabfirst = create_tab(tab_control, "Envelopes")
tabsecond = create_tab(tab_control, "Postal Traces")
tab1 = create_tab(tab_control, "Cure Letter")
tab2 = create_tab(tab_control, "Final Payment Letter")
tab3 = create_tab(tab_control, "VPA")
tab4 = create_tab(tab_control, "Payoff")
tab5 = create_tab(tab_control, "Checks")


tab_control.pack(expand=True, fill="both")

##########################
#
#   ENVELOPE TAB
#
###########################
def printReturn(filename):
    word.Documents.Open(filename)
    word.ActiveDocument.PrintOut()
    word.ActiveDocument.Close()


style = ttk.Style()
style.configure("White.TButton", background="white")

# Create a frame for the two side-by-side frames
dual_frame = tk.Frame(tabfirst)
dual_frame.pack(fill="both", pady=10)

# Create a frame for the 'Firm' section
firm_frame_e = tk.LabelFrame(dual_frame, text="Firm")
firm_frame_e.pack(side=tk.LEFT, padx=4, pady=2, fill="both")

firm_var_e = tk.StringVar()

had_radio_e = ttk.Radiobutton(firm_frame_e, text="HAD", variable=firm_var_e, value="HAD")
had_radio_e.pack(side=tk.LEFT, padx=3, pady=5)

nwltg_radio_e = ttk.Radiobutton(firm_frame_e, text="NWLTG", variable=firm_var_e, value="NWLTG")
nwltg_radio_e.pack(side=tk.LEFT, padx=3, pady=5)

noreturn_radio_e = ttk.Radiobutton(firm_frame_e, text="No Return Address", variable=firm_var_e, value="NORETURN")
noreturn_radio_e.pack(side=tk.LEFT, padx=3, pady=5)

# Create a frame for the 'Return Envelopes' section
return_frame = tk.LabelFrame(dual_frame, text="Return Envelopes")
return_frame.pack(side=tk.RIGHT, padx=3, pady=2, fill="both", expand=True,)

return_nwltg_button = tk.Button(return_frame, text="NWTLG", command=lambda:printWordDocument("F:\\DocumentCreator\\Envelopes\\nwreturn.docx"))
return_nwltg_button.pack(side=tk.LEFT, padx=5, pady=3)

return_had_button = tk.Button(return_frame, text="  HAD  ", command=lambda:printWordDocument("F:\\DocumentCreator\\Envelopes\\hadreturn.docx"))
return_had_button.pack(side=tk.LEFT, padx=5, pady=3)

# Service section
serv_var_e = tk.StringVar()
service_frame_e = tk.LabelFrame(tabfirst, text="Service Requested")
service_frame_e.pack(padx=3, pady=3, fill="both")

serv_var_e.set("Address")

address_radio_e = ttk.Radiobutton(service_frame_e, text="Address", variable=serv_var_e, value="Address", )
address_radio_e.pack(side=tk.LEFT, padx=5, pady=2)

change_radio_e = ttk.Radiobutton(service_frame_e, text="Change", variable=serv_var_e, value="Change")
change_radio_e.pack(side=tk.LEFT, padx=5, pady=2)

forwarding_radio_e = ttk.Radiobutton(service_frame_e, text="Forwarding", variable=serv_var_e, value="Forwarding")
forwarding_radio_e.pack(side=tk.LEFT, padx=5, pady=2)

return_radio_e = ttk.Radiobutton(service_frame_e, text="Return", variable=serv_var_e, value="Return")
return_radio_e.pack(side=tk.LEFT, padx=5, pady=2)

none_radio_e = ttk.Radiobutton(service_frame_e, text="None", variable=serv_var_e, value=" ")
none_radio_e.pack(side=tk.LEFT, padx=5, pady=2)

# Info section

info_frame_e = tk.LabelFrame(tabfirst, text="Information")
info_frame_e.pack(side=tk.TOP, fill="both")

# Name field
name_label_e = tk.Label(info_frame_e, text="Name:")
name_label_e.grid(row=0, column=0, sticky=tk.W, pady=2)
name_entry_e = tk.Entry(info_frame_e, width=45)
name_entry_e.focus_set()
name_entry_e.grid(row=0, column=1, sticky=tk.W, pady=2)

# Address field
addr_label_e = tk.Label(info_frame_e, text="Address:")
addr_label_e.grid(row=1, column=0, sticky=tk.W, pady=2)
addr_entry_e = tk.Entry(info_frame_e, width=45)
addr_entry_e.focus_set()
addr_entry_e.grid(row=1, column=1, sticky=tk.W, pady=2)

# CSZ field
csz_label_e = tk.Label(info_frame_e, text="City/State/Zip:  ")
csz_label_e.grid(row=2, column=0, sticky=tk.W, pady=2)
csz_entry_e = tk.Entry(info_frame_e, width=45)
csz_entry_e.focus_set()
csz_entry_e.grid(row=2, column=1, sticky=tk.W, pady=2)

# Notes field
notes_label_e = tk.Label(info_frame_e, text="Extra line:")
notes_label_e.grid(row=3, column=0, sticky=tk.W, pady=2)
notes_entry_e = tk.Entry(info_frame_e, width=45)
notes_entry_e.focus_set()
notes_entry_e.grid(row=3, column=1, sticky=tk.W, pady=2)

def printWordDocument(filename):
        word.Documents.Open(filename)
        word.ActiveDocument.PrintOut()
        word.ActiveDocument.Close()

def clearForms_e():
    name_entry_e.delete(0, 'end')
    addr_entry_e.delete(0, 'end')
    csz_entry_e.delete(0, 'end')
    notes_entry_e.delete(0, 'end')

def createEnvelope(f, service, name, addr, csz, recip, save):

    if f != 'NORETURN':
        doc_path = "F:\\DocumentCreator\\Envelopes\\envelope.docx"
    
    else:
        doc_path = "F:\\DocumentCreator\\Envelopes\\envelope_no_return_address.docx"
        print("ASDSAD")

    doc = DocxTemplate(doc_path)
    firm_name = attr = firm_addr = firm_csz = "   "
    
    if f == "HAD":
        firm_name = "HARRINGTON, ANDERSON & DeBLASIO"
        attr = "ATTORNEYS AT LAW"
        firm_addr = "P.O. Box 12669"
        firm_csz = "Portland, OR 97212"

    elif f == "NWLTG":
        firm_name = "Northwest Territory Law Group, P.C."
        attr = "Attorneys at Law"
        firm_addr = "3439 NE Sandy Blvd. #239"
        firm_csz = "Portland, OR 97232"          

    if service != " ":
        service = service.upper() + " SERVICE REQUESTED"

    elif service == " ":
        service = "ADDRESS SERVICE REQUESTED"

    context = {'Firm': firm_name, 'Attorney': attr, 'Firm_Addr': firm_addr, 'Firm_CSZ': firm_csz, 'Service': service,
               'Name': name, 'Addr': addr, 'CSV': csz, 'Recip': recip}
    if save == True:
        doc.render(context)
        save_path = filedialog.asksaveasfilename(defaultextension=".docx", filetypes=[("Word Files", "*.docx")], initialfile=name_entry_e.get() + ' - Envelope')
        doc.save(save_path)

    elif save == False:
        doc.render(context)
        doc.save("F:\\DocumentCreator\\Temp\\envelope.docx")
        printWordDocument("F:\\DocumentCreator\\Temp\\envelope.docx")
        os.remove("F:\\DocumentCreator\\Temp\\envelope.docx")


# Assign data from fields to variables
def submitFunction_e(save):
    firm = firm_var_e.get()
    service = serv_var_e.get()
    name = name_entry_e.get()
    addr = addr_entry_e.get()
    csz = csz_entry_e.get()
    recip = notes_entry_e.get()
    createEnvelope(firm, service, name, addr, csz, recip, save)

# Create a frame for the buttons
button_frame_e = tk.Frame(tabfirst)
button_frame_e.pack(pady=10)

# Create the Save button
save_button_e = tk.Button(button_frame_e, text="Save", width = 8, command=lambda:submitFunction_e(True))
save_button_e.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

# Create the Submit button
submit_button_e = tk.Button(button_frame_e, text="Print", command=lambda:submitFunction_e(False), width = 8)
submit_button_e.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

# Clear button
clear_button_e = tk.Button(button_frame_e, text="Clear", command=clearForms_e, width = 8)
clear_button_e.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

#############################
#
# POSTAL TRACE TAB
#
#############################

# Firm section
firm_frame_pt = tk.LabelFrame(tabsecond, text="Firm", padx=2, pady=2)
firm_frame_pt.pack(padx=3, pady=3, fill="both")

firm_var_pt = tk.StringVar()

had_radio_pt = tk.Radiobutton(firm_frame_pt, text="HAD", variable=firm_var_pt, value="HAD", tristatevalue=0)
had_radio_pt.grid(row=0, column=0, padx=5)

nwltg_radio_pt = tk.Radiobutton(firm_frame_pt, text="NWTLG", variable=firm_var_pt, value="NWTLG", tristatevalue=0)
nwltg_radio_pt.grid(row=0, column=1, padx=5)

# Information section
info_frame_pt = tk.LabelFrame(tabsecond, text="Information", padx=2, pady=2)
info_frame_pt.pack(padx=3, pady=3, fill="both")

# Name field
name_label_pt = tk.Label(info_frame_pt, text="Name:")
name_label_pt.grid(row=0, column=0, sticky=tk.W, pady=2)
name_entry_pt = tk.Entry(info_frame_pt, width=30)
name_entry_pt.focus_set()
name_entry_pt.grid(row=0, column=1, sticky=tk.W, pady=2)

# Address field
addr_label_pt = tk.Label(info_frame_pt, text="Address:")
addr_label_pt.grid(row=1, column=0, sticky=tk.W, pady=2)
addr_entry_pt = tk.Entry(info_frame_pt, width=30)
addr_entry_pt.focus_set()
addr_entry_pt.grid(row=1, column=1, sticky=tk.W, pady=2)

keepAddr = tk.BooleanVar()
keepCSZ = tk.BooleanVar()

addrCheckbox = ttk.Checkbutton(info_frame_pt, text="Don't clear", variable=keepAddr)
addrCheckbox.grid(row=1, column=2, sticky="w")  

cszCheckbox = ttk.Checkbutton(info_frame_pt, text="Don't clear", variable=keepCSZ)
cszCheckbox.grid(row=2, column=2, sticky="w") 

# CSZ field
csz_label_pt = tk.Label(info_frame_pt, text="City/State/Zip:")
csz_label_pt.grid(row=2, column=0, sticky=tk.W, pady=2)
csz_entry_pt = tk.Entry(info_frame_pt, width=30)
csz_entry_pt.focus_set()
csz_entry_pt.grid(row=2, column=1, sticky=tk.W, pady=2)

# Court field
court_label_pt = tk.Label(info_frame_pt, text="Court:")
court_label_pt.grid(row=3, column=0, sticky=tk.W, pady=2)
court_entry_pt = tk.Entry(info_frame_pt, width=30)
court_entry_pt.focus_set()
court_entry_pt.grid(row=3, column=1, sticky=tk.W, pady=2)

# Client field
client_label_pt = tk.Label(info_frame_pt, text="Plaintiff:")
client_label_pt.grid(row=4, column=0, sticky=tk.W, pady=2)
client_entry_pt = tk.Entry(info_frame_pt, width=30)
client_entry_pt.focus_set()
client_entry_pt.grid(row=4, column=1, sticky=tk.W, pady=2)

# Docket number field
num_label_pt = tk.Label(info_frame_pt, text="Reference number:")
num_label_pt.grid(row=5, column=0, sticky=tk.W, pady=2)
num_entry_pt = tk.Entry(info_frame_pt, width=30)
num_entry_pt.focus_set()
num_entry_pt.grid(row=5, column=1, sticky=tk.W, pady=2)

# PRINT SETTINGS SECTION
print_pt = tk.LabelFrame(tabsecond, text="Print Settings", padx=2, pady=2)
print_pt.pack(padx=3, pady=3, fill="both")

# Create variables to store the state of the checkboxes
return_var = tk.BooleanVar()
postmaster_var = tk.BooleanVar()

return_checkbox = ttk.Checkbutton(print_pt, text="No return envelope", variable=return_var)
return_checkbox.grid(row=0, column=0, sticky="w")  # Add 'sticky' option

postmaster_checkbox = ttk.Checkbutton(print_pt, text="No postmaster envelope", variable=postmaster_var)
postmaster_checkbox.grid(row=1, column=0, sticky="w")  # Add 'sticky' option

def createDocument_pt(f, n, a, z, c, num, cl, save):
    nup = n.upper()

    if num == "":
        num = "(tbd)"

    if f == "NWTLG":
        if return_var.get() and postmaster_var.get():  
            doc_path = "F:\\DocumentCreator\\PT\\nwpt_only.docx"
        elif return_var.get() and not postmaster_var.get():
            doc_path = "F:\\DocumentCreator\\PT\\nwpt_no_return.docx"
        elif not return_var.get() and postmaster_var.get():
            doc_path = "F:\\DocumentCreator\\PT\\nwpt_no_postmaster.docx"
        else:
            doc_path = "F:\\DocumentCreator\\PT\\nwpt.docx"

    elif f == "HAD":
        if return_var.get() and postmaster_var.get():  
            doc_path = "F:\\DocumentCreator\\PT\\hadpt_only.docx"
        elif return_var.get() and not postmaster_var.get():
            doc_path = "F:\\DocumentCreator\\PT\\hadpt_no_return.docx"
        elif not return_var.get() and postmaster_var.get():
            doc_path = "F:\\DocumentCreator\\PT\\hadpt_no_postmaster.docx"
        else:
            doc_path = "F:\\DocumentCreator\\PT\\hadpt.docx"        

    context = {'name': n, 'addr': a, 'csz': z, 'court': c, 'client': cl, 'name_up': nup,
               'num': num}
    
    doc = DocxTemplate(doc_path)

    if save == True:
        current_date = datetime.now().strftime("%m.%d.%y")
        file_name = name_entry_pt.get()
        file_name_with_date = f"{current_date} - {file_name} - Postal Trace"
        doc.render(context)
        save_path = filedialog.asksaveasfilename(defaultextension=".docx", filetypes=[("Word Files", "*.docx")], initialfile=file_name_with_date)
        doc.save(save_path)

    elif save == False:
        doc = DocxTemplate(doc_path)
        doc.render(context)
        doc.save("F:\\DocumentCreator\\Temp\\PTDoc.docx")
        printWordDocument("F:\\DocumentCreator\\Temp\\PTDoc.docx")  

def clearAddr():
    if not keepAddr.get():
        addr_entry_pt.delete(0, 'end')
    if not keepCSZ.get():
        csz_entry_pt.delete(0, 'end')

def submitFunction_pt(save):
    firm = firm_var_pt.get()
    name = name_entry_pt.get()
    addr = addr_entry_pt.get()
    csz = csz_entry_pt.get()
    court = court_entry_pt.get()
    num = num_entry_pt.get()
    client = client_entry_pt.get()
    createDocument_pt(firm, name, addr, csz, court, num, client, save)
    #os.remove("C:\\Users\\mark\\Automations\\pt_generated.docx")
    clearAddr()

def clearForms_pt():
    name_entry_pt.delete(0, 'end')
    if not keepAddr.get():
        addr_entry_pt.delete(0, 'end')
    if not keepCSZ.get():
        csz_entry_pt.delete(0, 'end')
    court_entry_pt.delete(0, 'end')
    client_entry_pt.delete(0, 'end')
    num_entry_pt.delete(0, 'end')

# Create a frame for the buttons
button_frame_pt = tk.Frame(tabsecond)
button_frame_pt.pack(pady=10)

# Create the Save button
save_button_pt = tk.Button(button_frame_pt, text="Save", width = 8, command=lambda:submitFunction_pt(True))
save_button_pt.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

# Create the Submit button
submit_button_pt = tk.Button(button_frame_pt, text="Print", command=lambda:submitFunction_pt(False), width = 8)
submit_button_pt.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

# Clear button
clear_button_pt = tk.Button(button_frame_pt, text="Clear", command=clearForms_pt, width = 8)
clear_button_pt.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

##########################
#
#   CURE TAB
#
###########################
# Firm section
firm_frame = tk.LabelFrame(tab1, text="Firm", padx=2, pady=2)
firm_frame.pack(padx=3, pady=3, fill="both")

firm_var = tk.StringVar()

had_radio = tk.Radiobutton(firm_frame, text="HAD", variable=firm_var, value="HAD", tristatevalue=0)
had_radio.pack(side=tk.LEFT, padx=5)

nwltg_radio = tk.Radiobutton(firm_frame, text="NWTLG", variable=firm_var, value="NWTLG", tristatevalue=0)
nwltg_radio.pack(side=tk.LEFT, padx=5)

# Info section

info_frame = tk.LabelFrame(tab1, text="Information", padx=2, pady=2)
info_frame.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Name field
name_label = tk.Label(info_frame, text="Name:")
name_label.grid(row=0, column=0, sticky=tk.W, pady=2)
name_entry = tk.Entry(info_frame, width=30)
name_entry.focus_set()
name_entry.grid(row=0, column=1, sticky=tk.W, pady=2)

# Address field
addr_label = tk.Label(info_frame, text="Address:")
addr_label.grid(row=1, column=0, sticky=tk.W, pady=2)
addr_entry = tk.Entry(info_frame, width=30)
addr_entry.focus_set()
addr_entry.grid(row=1, column=1, sticky=tk.W, pady=2)

# CSZ field
csz_label = tk.Label(info_frame, text="City/State/Zip:")
csz_label.grid(row=2, column=0, sticky=tk.W, pady=2)
csz_entry = tk.Entry(info_frame, width=30)
csz_entry.focus_set()
csz_entry.grid(row=2, column=1, sticky=tk.W, pady=2)

# Email field
email_label = tk.Label(info_frame, text="Email:")
email_label.grid(row=3, column=0, sticky=tk.W, pady=2)
email_entry = tk.Entry(info_frame, width=30)
email_entry.focus_set()
email_entry.grid(row=3, column=1, sticky=tk.W, pady=2)

# Plaintiff field
plaintiff_label = tk.Label(info_frame, text="Plaintiff:")
plaintiff_label.grid(row=0, column=2, sticky=tk.W, pady=2)
plaintiff_entry = tk.Entry(info_frame, width=30)
plaintiff_entry.focus_set()
plaintiff_entry.grid(row=0, column=3, sticky=tk.W, pady=2)

# Court field
court_label = tk.Label(info_frame, text="Court:")
court_label.grid(row=1, column=2, sticky=tk.W, pady=2)
court_entry = tk.Entry(info_frame, width=30)
court_entry.focus_set()
court_entry.grid(row=1, column=3, sticky=tk.W, pady=2)

# Case number field
case_label = tk.Label(info_frame, text="Case No:")
case_label.grid(row=2, column=2, sticky=tk.W, pady=2)
case_entry = tk.Entry(info_frame, width=20)
case_entry.focus_set()
case_entry.grid(row=2, column=3, sticky=tk.W, pady=2)

# Jxt amount
jxt_label = tk.Label(info_frame, text="Jmt amount:")
jxt_label.grid(row=3, column=2, sticky=tk.W, pady=2)
jxt_entry = tk.Entry(info_frame, width=10)
jxt_entry.focus_set()
jxt_entry.grid(row=3, column=3, sticky=tk.W, pady=2)

# Amount paid
amt_label = tk.Label(info_frame, text="Less:")
amt_label.grid(row=4, column=2, sticky=tk.W, pady=2)
amt_entry = tk.Entry(info_frame, width=10)
amt_entry.focus_set()
amt_entry.grid(row=4, column=3, sticky=tk.W, pady=2)

# Payments frame
payments_frame = tk.LabelFrame(tab1, text="Payments", padx=2, pady=2)
payments_frame.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Last payment amount
lpayamt_label = tk.Label(payments_frame, text="Last Payment:")
lpayamt_label.grid(row=0, column=0, sticky=tk.W, pady=2)
lpayamt_entry = tk.Entry(payments_frame, width=10)
lpayamt_entry.focus_set()
lpayamt_entry.grid(row=0, column=1, sticky=tk.W, pady=2)

# Min payment
min_label = tk.Label(payments_frame, text="Min Payment:")
min_label.grid(row=1, column=0, sticky=tk.W, pady=2)
min_entry = tk.Entry(payments_frame, width=10)
min_entry.focus_set()
min_entry.grid(row=1, column=1, sticky=tk.W, pady=2)

# Day due
day_label = tk.Label(payments_frame, text="Day due:")
day_label.grid(row=2, column=0, sticky=tk.W, pady=2)
day_entry = tk.Entry(payments_frame, width=10)
day_entry.focus_set()
day_entry.grid(row=2, column=1, sticky=tk.W, pady=2)

# End of month text
eom_text = "(Leave blank if payment is due before the end of the month)"
eom_label = tk.Label(payments_frame, text=eom_text)
eom_label.place(x=177, y=66)

def toggle_resume_entry():
    if resumePay.get():
        resume_entry.configure(state='normal')
    else:
        resume_entry.configure(state='disabled')

# Add a checkbox with variable resumePay
resumePay = tk.BooleanVar()
resumePay.set(False)
resume_checkbox = ttk.Checkbutton(payments_frame, text="Resume by:", variable=resumePay, command=toggle_resume_entry)
resume_checkbox.grid(row=3, column=0, sticky=tk.W, pady=2)

resume_entry = DateEntry(payments_frame, width=12, background="white", foreground="black", state='disabled')
resume_entry.grid(row=3, column=1, sticky=tk.W, pady=2)

# Last payment
lpay_label = tk.Label(payments_frame, text="on")
lpay_label.grid(row=0, column=2, sticky=tk.W, pady=2)
lpay_entry = DateEntry(payments_frame, width=12, background="white", foreground="black")
lpay_entry.grid(row=0, column=3, sticky=tk.W, pady=2)

# Applied to
lremit_label = tk.Label(payments_frame, text="Remit Applied To:")
lremit_label.grid(row=0, column=4, sticky=tk.W, pady=2)
lremit_entry = DateEntry(payments_frame, width=12, background="white", foreground="black")
lremit_entry.grid(row=0, column=5, sticky=tk.W, pady=2)

def open_file_dialog():
    file_path = filedialog.askopenfilename()
    if file_path:
        file_label.config(text="Selected File: " + file_path)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        # Remove brackets and commas
        data = [item for sublist in data for item in sublist]
        insertOutlineData(data)

    else:
        file_label.config(text="No file selected")

# Upload Outline Frame
outline_frame = tk.LabelFrame(tab1, text="Upload Outline", padx=2, pady=2)
outline_frame.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Upload Button
upload_button = tk.Button(outline_frame, text="Upload", command=open_file_dialog)
upload_button.grid(row=0, column=0, padx=5, pady=5)

# Label to display selected file name
file_label = tk.Label(outline_frame, text="No file selected")
file_label.grid(row=0, column=1, padx=5, pady=5)

def printWordDocument(filename):
    word.Documents.Open(filename)
    word.ActiveDocument.PrintOut()
    time.sleep(2)
    word.ActiveDocument.Close()

def removeDay(date):
    # Split the date string into month, day, and year parts
    parts = date.split()
    month = parts[0]
    year = parts[2]

    # Format the new date as "Month Year"
    new_date_formatted = f"{month} {year}"

    return new_date_formatted

def processDate(day, day2):
    # Get the current date
    current_date = datetime.now()

    # Convert the day to an integer
    day = int(day)
    print(current_date.day)
    # Check if the current day matches the specified day
    if day >= current_date.day or day2 == 0:
        # Subtract a month and set the day to the specified day
        new_date = current_date.replace(month=current_date.month - 1, day=day)
    else:
        # Create a new date object with the specified day
        new_date = date(current_date.year, current_date.month, day)

    # Format the new date as "Month Day, Year"
    new_date_formatted = new_date.strftime("%B %d, %Y")

    return new_date_formatted

def addMonth(date_string):
    # Convert date string to datetime object
    date_format = "%m/%d/%y"
    date_object = datetime.strptime(date_string, date_format)

    # Calculate the new month and year
    current_month = date_object.month
    current_year = date_object.year
    new_month = current_month + 1
    new_year = current_year

    # Adjust the month and year if December is reached
    if new_month == 13:
        new_month = 1
        new_year += 1

    # Add a month to the date while keeping the day the same
    new_date = date_object + relativedelta(months=new_month - date_object.month, years=new_year - date_object.year)

    # Format the new date as "Month Day, Year"
    new_date_formatted = new_date.strftime("%B %d, %Y")

    return new_date_formatted

def calcMonths(date1, date2):
    # Convert date strings to datetime objects
    date_format = "%m/%d/%y"
    datetime1 = datetime.strptime(date1, date_format)
    datetime2 = datetime.strptime(date2, date_format)

    # Calculate the difference in months
    month_difference = (datetime2.year - datetime1.year) * 12 + (datetime2.month - datetime1.month)

    return month_difference

def suffix(day):
    number = int(day)
    suffix = "th"

    if number % 10 == 1 and number % 100 != 11:
        suffix = "st"
    elif number % 10 == 2 and number % 100 != 12:
        suffix = "nd"
    elif number % 10 == 3 and number % 100 != 13:
        suffix = "rd"

    return suffix

def nextMonth(day):
    current_date = datetime.today().date()
    next_month = current_date + relativedelta(months=1)
    next_month_date = next_month.replace(day=int(day))
    return next_month_date.strftime("%B %d, %Y")

def genFile(name, type):
    current_date = datetime.today().date().strftime("%m.%d.%y")
    file_name = f"{current_date} - {name} - {type}.docx"
    return file_name

def dateConvert(date_string):
    date_obj = datetime.strptime(date_string, '%m/%d/%y')
    formatted_date = date_obj.strftime('%B %d, %Y')
    return formatted_date

def floatConvert(value):
    try:
        float_value = float(value.replace(',', ''))
        return round(float_value, 2)
    except ValueError:
        return None
    
def convertDecimal(value):
    try:
        decimal_value = Decimal(str(value))
        return decimal_value.quantize(Decimal('0.00'))
    except (ValueError, TypeError):
        return None

def dateConvertDayRm(date_str):
    # Parse the date string into a datetime object
    date_obj = datetime.strptime(date_str, "%m/%d/%y")

    # Format the date as "Month, Day, Year"
    formatted_date = date_obj.strftime("%B %d, %Y")

    return formatted_date

def removeDayRemit(date_str):
    parts = date_str.split(",")  # Split the date string by comma
    month_day = parts[0].strip()  # Extract the month and day part and remove leading/trailing spaces

    # Split the month and day by space and take only the first part (month)
    month = month_day.split()[0]
    year = parts[1].strip()  # Extract the year part and remove leading/trailing spaces

    return f"{month} {year}"

def formatNumberWithCommas(number_str):
    # Set the appropriate locale
    locale.setlocale(locale.LC_ALL, '')

    # Convert the number string to float
    number = float(number_str)

    # Format the number with commas
    formatted_number = locale.format_string("%0.2f", number, grouping=True)

    return str(formatted_number)

def removeFirstYear(date_range):
    parts = date_range.split()  # Split the string by space
    start_year = parts[1]  # Get the year from the start date
    end_year = parts[-1]  # Get the year from the end date

    # Check if the years match
    if start_year == end_year:
        parts.pop(1)  # Remove the year from the start date

    return ' '.join(parts)  # Join the parts back into a string

def is_day_less_than_current(day):
    # Get the current day as an integer
    current_day = date.today().day

    # Convert the 'day' variable to an integer
    input_day = int(day)

    # Compare the input day with the current day
    return input_day <= current_day

def get_current_date_with_day(day):
    # Get the current date
    current_date = date.today()

    # Get the current month and year
    current_month = current_date.strftime('%B')
    current_year = current_date.year

    # Return the string with the current month, the 'day' variable, and the current year
    return f"{current_month} {day}, {current_year}"

def isConsecutive(missed_day_start, missed_day_finish, day2):
    if day2 != 0:
        start_month, _, _ = missed_day_start.split(' ')
        finish_month, _, _ = missed_day_finish.split(' ')

    else:
        start_month, _ = missed_day_start.split(' ')
        finish_month, _ = missed_day_finish.split(' ')

    months = {
        'January': 1, 'February': 2, 'March': 3, 'April': 4,
        'May': 5, 'June': 6, 'July': 7, 'August': 8,
        'September': 9, 'October': 10, 'November': 11, 'December': 12
    }

    return months[finish_month] - months[start_month] == 1

def formatResume(date):
    # Convert input string to a datetime object
    date_obj = datetime.strptime(date, '%m/%d/%y')

    # Format the date as "Month Day, Year"
    formatted_date = date_obj.strftime('%B %d, %Y')

    return formatted_date

def createDocument(firm, name, addr, csz, email, court, case, jxt, amt, min_en, lamt, day, lpay, lremit, plaintiff, save, resumePay, resumeDate):
    day2 = day
    if day == "":
        min_day = "before the end"
        day = 1
        day2 = 0
    else:
        min_day = "on or before the " + day + suffix(day)

    current_date = datetime.now().date()
    # Format date as "mm/dd/YYYY"
    datef = current_date.strftime("%m/%d/%y")

    # Months between last remit and this month
    months = calcMonths(lremit, datef)
    if day2 == 0:
        months = months - 1

    if months > 1:
        missed_day_start = addMonth(lremit)
        missed_day_finish = processDate(day, day2)
        if day2 == 0:
            missed_day_start = removeDay(missed_day_start)
            missed_day_finish = removeDay(missed_day_finish)

        if isConsecutive(missed_day_start, missed_day_finish, day2):
            missed =  removeFirstYear(missed_day_start + ", and on " + missed_day_finish)
            missed_pay = "payments that were due on"
        else:
            missed =  removeFirstYear(missed_day_start + ", to " + missed_day_finish)
            missed_pay = "payments that were due from"
        print("MISSED")
        print(missed)
    
    else:
        missed_pay = "payment due on"
        missed_day_start = addMonth(lremit)
        missed = missed_day_start

    if day2 != 0:
        if is_day_less_than_current(day) == True:
            next = nextMonth(day)    
        else:
            print(day)
            next = get_current_date_with_day(day)
        next_due = "on or before"
        lremit = dateConvert(lremit)
    elif day2 == 0: 
        next = removeDay(addMonth(datef))
        next_due = "before the end of"
        print(dateConvertDayRm(lremit))
        lremit = removeDayRemit(dateConvertDayRm(lremit))
        print(lremit)

    if amt == "":
        amt = '0'

    min_payment = convertDecimal(float(min_en))
    min_bal = formatNumberWithCommas(convertDecimal(float(min_payment * months)))
    lamt = convertDecimal(float(lamt))
    due = formatNumberWithCommas(convertDecimal(floatConvert(jxt)) - convertDecimal(floatConvert(amt)))
    min_en = formatNumberWithCommas(min_en)

    lpay = dateConvert(lpay) + ","

    if email != "" and not resumePay:
        email = "*Sent via email: " + email
        if firm == 'NWTLG':
            doc_path = "F:\\DocumentCreator\\Cure\\nwcure_email.docx"
        elif firm == 'HAD':
            doc_path = "F:\\DocumentCreator\\Cure\\hadcure_email.docx"

    elif email == "" and not resumePay:
        if firm == 'NWTLG':
            doc_path = "F:\\DocumentCreator\\Cure\\nwcure.docx"
        else:
            doc_path = "F:\\DocumentCreator\\Cure\\hadcure.docx"

    elif email != "" and resumePay:
        email = "*Sent via email: " + email
        if firm == 'NWTLG':
            doc_path = "F:\\DocumentCreator\\Cure\\nwresume_email.docx"
        elif firm == 'HAD':
            doc_path = "F:\\DocumentCreator\\Cure\\hadresume_email.docx"

    elif email == "" and resumePay:
        if firm == 'NWTLG':
            doc_path = "F:\\DocumentCreator\\Cure\\nwresume.docx"
        elif firm == 'HAD':
            doc_path = "F:\\DocumentCreator\\Cure\\hadresume.docx"

    doc = DocxTemplate(doc_path)

    context = {'name': name, 'addr': addr, 'csz': csz, 'email': email, 'court': court, 'case': case, 
               'bal': due, 'last_pay': lamt, 'last_date': lpay, 'last_remit': lremit,
               'min': min_en, 'missed_pay': missed_pay, 'missed_dates': missed, 'm_bal': min_bal, 
               'next_pay': next, 'next_due': next_due, 'min_day': min_day, 'plaintiff': plaintiff, 'init': initials_input.get(), 'resume_date': formatResume(resumeDate)}

    if save == True:
        current_date = datetime.now().strftime("%m.%d.%y")
        file_name = name_entry.get()
        if not resumePay:
            file_name_with_date = f"{current_date} - {file_name} - Cure Letter"
        else:
            file_name_with_date = f"{current_date} - {file_name} - Payment Resumption Letter"
        doc.render(context)
        save_path = filedialog.asksaveasfilename(defaultextension=".docx", filetypes=[("Word Files", "*.docx")], initialfile=file_name_with_date)
        doc.save(save_path)

    elif save == False:
        doc.render(context)
        doc.save("F:\\DocumentCreator\\Temp\\cure.docx")
        printWordDocument("F:\\DocumentCreator\\Temp\\cure.docx")
        os.remove("F:\\DocumentCreator\\Temp\\cure.docx")

def modify_string(s):
    if s.strip().lower().startswith("onemain"):
        index = s.lower().find("llc")
        if index != -1:
            s = s[:index + 3]  # Include "LLC" in the result

            words = s.split()  # Split the string into words
            modified_words = []
            inside_onemain = False
            
            for word in words:
                if word.lower() == "onemain":
                    modified_words.append("OneMain")
                    inside_onemain = True
                elif word.lower() == "llc":
                    modified_words.append(word)
                    inside_onemain = False
                elif inside_onemain:
                    modified_words.append(word.capitalize())
                else:
                    modified_words.append(word)
            
            return " ".join(modified_words)
    
def clean_number(number_str):
    cleaned_str = number_str.replace("$", "").replace(",", "")
    return cleaned_str

def insertOutlineData(outlineData):
    if outlineData[0] == 'NWTLG':
        firm_var.set("NWTLG")
    elif outlineData[0] == 'HAD':
        firm_var.set("HAD")

    name_entry.delete(0, 'end')
    name_entry.insert(0, outlineData[1])
    addr_entry.delete(0, 'end')
    addr_entry.insert(0, outlineData[2])
    csz_entry.delete(0, 'end')
    csz_entry.insert(0, outlineData[3] + ', '+ outlineData[4] + ' ' + outlineData[5])
    email_entry.delete(0, 'end')
    email_entry.insert(0, outlineData[6])
    plaintiff_entry.delete(0, 'end')
    plaintiff_entry.insert(0, modify_string(outlineData[7]))
    court_entry.delete(0, 'end')
    court_entry.insert(0, outlineData[8])
    case_entry.delete(0, 'end')
    case_entry.insert(0, outlineData[9])
    jxt_entry.delete(0, 'end')
    jxt_entry.insert(0, clean_number(outlineData[10]))
    min_entry.delete(0, 'end')
    min_entry.insert(0, clean_number(outlineData[11]))
    day_entry.delete(0, 'end')
    day_entry.insert(0, outlineData[12])

def clearFunction():
    name_entry.delete(0, 'end')
    addr_entry.delete(0, 'end')
    csz_entry.delete(0, 'end')
    email_entry.delete(0, 'end')
    court_entry.delete(0, 'end')
    case_entry.delete(0, 'end')
    jxt_entry.delete(0, 'end')
    amt_entry.delete(0, 'end')
    min_entry.delete(0, 'end')
    lpayamt_entry.delete(0, 'end')
    day_entry.delete(0, 'end')
    lpay_entry.delete(0, 'end')
    lremit_entry.delete(0, 'end')
    plaintiff_entry.delete(0, 'end')
    file_label.config(text="No file selected")
    resume_entry.delete(0, 'end')

def submitFunction(save):
    firm = firm_var.get()
    name = name_entry.get()
    addr = addr_entry.get()
    csz = csz_entry.get()
    email = email_entry.get()
    court = court_entry.get()
    case = case_entry.get()
    jxt = jxt_entry.get()
    amt = amt_entry.get()
    min_en = min_entry.get()
    lamt = lpayamt_entry.get()
    day = day_entry.get()
    lpay = lpay_entry.get()
    lremit = lremit_entry.get()
    plaintiff = plaintiff_entry.get()
    resumeBool = resumePay.get()
    resumeDate = resume_entry.get()

    createDocument(firm, name, addr, csz, email, court, case, jxt, amt, min_en, lamt, day, lpay, lremit, plaintiff, save, resumeBool, resumeDate)

# Create a frame for the buttons
button_frame = tk.Frame(tab1)
button_frame.pack(pady=10)

# Create the Save button
save_button_cure = tk.Button(button_frame, text="Save", width = 8, command=lambda:submitFunction(True))
save_button_cure.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

# Create the Print button
submit_button = tk.Button(button_frame, text="Print", command=lambda:submitFunction(False), width = 8)
submit_button.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

# Create the Clear button
clear_button = tk.Button(button_frame, text="Clear", command=clearFunction, width = 8)
clear_button.pack(side=tk.LEFT, anchor="w", padx=(5, 0))


##########################
#
#   FINAL PAYMENT TAB
#
###########################
# Firm section
firm_frame = tk.LabelFrame(tab2, text="Firm", padx=2, pady=2)
firm_frame.pack(padx=3, pady=3, fill="both")

firm_var_t2 = tk.StringVar()

had_radio_t2 = tk.Radiobutton(firm_frame, text="HAD", variable=firm_var_t2, value="HAD", tristatevalue=0)
had_radio_t2.pack(side=tk.LEFT, padx=5)

nwltg_radio_t2 = tk.Radiobutton(firm_frame, text="NWTLG", variable=firm_var_t2, value="NWTLG", tristatevalue=0)
nwltg_radio_t2.pack(side=tk.LEFT, padx=5)

# Info section

info_frame_t2 = tk.LabelFrame(tab2, text="Information", padx=2, pady=2)
info_frame_t2.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Name field
name_label_t2 = tk.Label(info_frame_t2, text="Name:")
name_label_t2.grid(row=0, column=0, sticky=tk.W, pady=2)
name_entry_t2 = tk.Entry(info_frame_t2, width=30)
name_entry_t2.focus_set()
name_entry_t2.grid(row=0, column=1, sticky=tk.W, pady=2)

# Address field
addr_label_t2 = tk.Label(info_frame_t2, text="Address:")
addr_label_t2.grid(row=1, column=0, sticky=tk.W, pady=2)
addr_entry_t2 = tk.Entry(info_frame_t2, width=30)
addr_entry_t2.focus_set()
addr_entry_t2.grid(row=1, column=1, sticky=tk.W, pady=2)

# CSZ field
csz_label_t2 = tk.Label(info_frame_t2, text="City/State/Zip:")
csz_label_t2.grid(row=2, column=0, sticky=tk.W, pady=2)
csz_entry_t2 = tk.Entry(info_frame_t2, width=30)
csz_entry_t2.focus_set()
csz_entry_t2.grid(row=2, column=1, sticky=tk.W, pady=2)

# Email field
email_label_t2 = tk.Label(info_frame_t2, text="Email:")
email_label_t2.grid(row=3, column=0, sticky=tk.W, pady=2)
email_entry_t2 = tk.Entry(info_frame_t2, width=30)
email_entry_t2.focus_set()
email_entry_t2.grid(row=3, column=1, sticky=tk.W, pady=2)

# Client field
client_label_t2 = tk.Label(info_frame_t2, text="Plaintiff:")
client_label_t2.grid(row=0, column=2, sticky=tk.W, pady=2)
client_entry_t2 = tk.Entry(info_frame_t2, width=30)
client_entry_t2.focus_set()
client_entry_t2.grid(row=0, column=3, sticky=tk.W, pady=2)

# Court field
court_label_t2 = tk.Label(info_frame_t2, text="Court:")
court_label_t2.grid(row=1, column=2, sticky=tk.W, pady=2)
court_entry_t2 = tk.Entry(info_frame_t2, width=30)
court_entry_t2.focus_set()
court_entry_t2.grid(row=1, column=3, sticky=tk.W, pady=2)

# Case number field
case_label_t2 = tk.Label(info_frame_t2, text="Case No:")
case_label_t2.grid(row=2, column=2, sticky=tk.W, pady=2)
case_entry_t2 = tk.Entry(info_frame_t2, width=20)
case_entry_t2.focus_set()
case_entry_t2.grid(row=2, column=3, sticky=tk.W, pady=2)

# Balance frame
bal_frame = tk.LabelFrame(tab2, text="Remaining balance", padx=2, pady=2)
bal_frame.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Remaining amount
rem_label = tk.Label(bal_frame, text="Balance Owing:")
rem_label.grid(row=0, column=0, sticky=tk.W, pady=2)
rem_entry = tk.Entry(bal_frame, width=10)
rem_entry.focus_set()
rem_entry.grid(row=0, column=1, sticky=tk.W, pady=2)

# Last payment
lday_label = tk.Label(bal_frame, text="on")
lday_label.grid(row=0, column=2, sticky=tk.W, pady=2)
lday_entry = DateEntry(bal_frame, width=12, background="white", foreground="black")
lday_entry.grid(row=0, column=3, sticky=tk.W, pady=2)

def open_file_dialog_t2():
    file_path = filedialog.askopenfilename()
    if file_path:
        file_label_t2.config(text="Selected File: " + file_path)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        # Remove brackets and commas
        data = [item for sublist in data for item in sublist]
        insertOutlineData_t2(data)

    else:
        file_label_t2.config(text="No file selected")

# Upload Outline Frame
outline_frame_t2 = tk.LabelFrame(tab2, text="Upload Outline", padx=2, pady=2)
outline_frame_t2.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Upload Button
upload_button_t2 = tk.Button(outline_frame_t2, text="Upload", command=open_file_dialog_t2)
upload_button_t2.grid(row=0, column=0, padx=5, pady=5)

# Label to display selected file name
file_label_t2 = tk.Label(outline_frame_t2, text="No file selected")
file_label_t2.grid(row=0, column=1, padx=5, pady=5)

def split(number):
    number = float(number)
    # Multiply the number by 100 to remove decimal places
    multiplied_number = number * 100
    
    # Check if the multiplied number is an even integer
    if int(multiplied_number) % 2 == 0:
        return True
    
    return False

def splitTwo(number):
    payment1 = round(number / 2, 2)
    payment2 = round(number - payment1, 2)
    return payment1, payment2

def convertDate_t2(date_string):
    # Split the date string into month, day, and year
    month, day, year = date_string.split('/')
    
    # Create a list of month names
    month_names = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    
    # Convert the month, day, and year to integers
    month = int(month)
    day = int(day)
    
    # Handle the year format
    if len(year) == 2:
        year = '20' + year
    else:
        year = str(year)
    
    # Format the date as "Month Day, Year"
    formatted_date = f"{month_names[month - 1]} {day}, {year}"
    
    return formatted_date

def createDocument_t2(firm, name, addr, csz, email, client, court, case, rem, lday):

    if split(rem):
        mpay = 'equal monthly payments'
        pay = "${:.2f}".format(float(rem)/2) + ' each'
    else:
        mpay = 'monthly payments'
        payment1, payment2 = splitTwo(float(rem))
        pay = "${:.2f}".format(payment1) + ' and ' + "${:.2f}".format(payment2) + ' each'

    if email != "":
        email = "*Sent via email: " + email
        if firm == 'NWTLG':
            doc_path = "F:\\DocumentCreator\\FinalPay\\nwfp_email.docx"
        elif firm == 'HAD':
            doc_path = "F:\\DocumentCreator\\FinalPay\\hadfp_email.docx"

    elif email == "":
        if firm == 'NWTLG':
            doc_path = "F:\\DocumentCreator\\FinalPay\\nwfp.docx"
        else:
            doc_path = "F:\\DocumentCreator\\FinalPay\\hadfp.docx"

    doc = DocxTemplate(doc_path)

    context = {'name': name, 'addr': addr, 'csz': csz, 'email': email, 'client': client, 'court': court, 'case': case,
               'bal': "${:.2f}".format(float(rem)), 'month_pay': mpay, 'pay': pay, 'end_date': convertDate_t2(lday), 'init': initials_input.get()}  

    doc.render(context)
    current_date = datetime.now().strftime("%m.%d.%y")
    file_name = name_entry_t2.get()
    file_name_with_date = f"{current_date} - {file_name} - Final Payment Letter"
    doc.render(context)
    save_path = filedialog.asksaveasfilename(defaultextension=".docx", filetypes=[("Word Files", "*.docx")], initialfile=file_name_with_date)
    doc.save(save_path)

def clearFunction_t2():
    name_entry_t2.delete(0, 'end')
    addr_entry_t2.delete(0, 'end')
    csz_entry_t2.delete(0, 'end')
    email_entry_t2.delete(0, 'end')
    client_entry_t2.delete(0, 'end')
    court_entry_t2.delete(0, 'end')
    case_entry_t2.delete(0, 'end')
    rem_entry.delete(0, 'end')
    lday_entry.delete(0, 'end')

def insertOutlineData_t2(outlineData):
    if outlineData[0] == 'NWTLG':
        firm_var_t2.set("NWTLG")
    elif outlineData[0] == 'HAD':
        firm_var_t2.set("HAD")

    name_entry_t2.delete(0, 'end')
    name_entry_t2.insert(0, outlineData[1])
    addr_entry_t2.delete(0, 'end')
    addr_entry_t2.insert(0, outlineData[2])
    csz_entry_t2.delete(0, 'end')
    csz_entry_t2.insert(0, outlineData[3] + ', '+ outlineData[4] + ' ' + outlineData[5])
    email_entry_t2.delete(0, 'end')
    email_entry_t2.insert(0, outlineData[6])
    client_entry_t2.delete(0, 'end')
    client_entry_t2.insert(0, modify_string(outlineData[7]))
    court_entry_t2.delete(0, 'end')
    court_entry_t2.insert(0, outlineData[8])
    case_entry_t2.delete(0, 'end')
    case_entry_t2.insert(0, outlineData[9])

def submitFunction_t2():
    firm = firm_var_t2.get()
    name = name_entry_t2.get()
    addr = addr_entry_t2.get()
    csz = csz_entry_t2.get()
    email = email_entry_t2.get()
    client = client_entry_t2.get()
    court = court_entry_t2.get()
    case = case_entry_t2.get()
    rem = rem_entry.get()
    lday = lday_entry.get()

    createDocument_t2(firm, name, addr, csz, email, client, court, case, rem, lday)

# Create a frame for the buttons
button_frame_t2 = tk.Frame(tab2)
button_frame_t2.pack(pady=10)

# Create the Save button
submit_button_t2 = tk.Button(button_frame_t2, text="Save", width = 8, command=submitFunction_t2)
submit_button_t2.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

# Create the Clear button
clear_button_t2 = tk.Button(button_frame_t2, text="Clear", width = 8, command=clearFunction_t2)
clear_button_t2.pack(side=tk.LEFT, anchor="w", padx=(5, 0))



##########################
#
#   VPA TAB
#
###########################

# Firm section
firm_frame = tk.LabelFrame(tab3, text="Firm", padx=2, pady=2)
firm_frame.pack(padx=3, pady=3, fill="both")

firm_var_t3 = tk.StringVar()

had_radio_t3 = tk.Radiobutton(firm_frame, text="HAD", variable=firm_var_t3, value="HAD", tristatevalue=0)
had_radio_t3.pack(side=tk.LEFT, padx=5)

nwltg_radio_t3 = tk.Radiobutton(firm_frame, text="NWTLG", variable=firm_var_t3, value="NWTLG", tristatevalue=0)
nwltg_radio_t3.pack(side=tk.LEFT, padx=5)

# Info section

info_frame_t3 = tk.LabelFrame(tab3, text="Information", padx=2, pady=2)
info_frame_t3.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Name field
name_label_t3 = tk.Label(info_frame_t3, text="Name:")
name_label_t3.grid(row=0, column=0, sticky=tk.W, pady=2)
name_entry_t3 = tk.Entry(info_frame_t3, width=30)
name_entry_t3.focus_set()
name_entry_t3.grid(row=0, column=1, sticky=tk.W, pady=2)

# Address field
addr_label_t3 = tk.Label(info_frame_t3, text="Address:")
addr_label_t3.grid(row=1, column=0, sticky=tk.W, pady=2)
addr_entry_t3 = tk.Entry(info_frame_t3, width=30)
addr_entry_t3.focus_set()
addr_entry_t3.grid(row=1, column=1, sticky=tk.W, pady=2)

# CSZ field
csz_label_t3 = tk.Label(info_frame_t3, text="City/State/Zip:")
csz_label_t3.grid(row=2, column=0, sticky=tk.W, pady=2)
csz_entry_t3 = tk.Entry(info_frame_t3, width=30)
csz_entry_t3.focus_set()
csz_entry_t3.grid(row=2, column=1, sticky=tk.W, pady=2)

# Email field
email_label_t3 = tk.Label(info_frame_t3, text="Email:")
email_label_t3.grid(row=3, column=0, sticky=tk.W, pady=2)
email_entry_t3 = tk.Entry(info_frame_t3, width=30)
email_entry_t3.focus_set()
email_entry_t3.grid(row=3, column=1, sticky=tk.W, pady=2)

# Client field
client_label_t3 = tk.Label(info_frame_t3, text="Plaintiff:")
client_label_t3.grid(row=0, column=2, sticky=tk.W, pady=2)
client_entry_t3 = tk.Entry(info_frame_t3, width=30)
client_entry_t3.focus_set()
client_entry_t3.grid(row=0, column=3, sticky=tk.W, pady=2)

# Court field
court_label_t3 = tk.Label(info_frame_t3, text="Court:")
court_label_t3.grid(row=1, column=2, sticky=tk.W, pady=2)
court_entry_t3 = tk.Entry(info_frame_t3, width=30)
court_entry_t3.focus_set()
court_entry_t3.grid(row=1, column=3, sticky=tk.W, pady=2)

# Case number field
case_label_t3 = tk.Label(info_frame_t3, text="Case No:")
case_label_t3.grid(row=2, column=2, sticky=tk.W, pady=2)
case_entry_t3 = tk.Entry(info_frame_t3, width=20)
case_entry_t3.focus_set()
case_entry_t3.grid(row=2, column=3, sticky=tk.W, pady=2)

# Balance frame
bal_frame_t3 = tk.LabelFrame(tab3, text="Payment", padx=2, pady=2)
bal_frame_t3.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Min payment
min_label_t3 = tk.Label(bal_frame_t3, text="Min payment:")
min_label_t3.grid(row=0, column=0, sticky=tk.W, pady=2)
min_entry_t3 = tk.Entry(bal_frame_t3, width=10)
min_entry_t3.focus_set()
min_entry_t3.grid(row=0, column=1, sticky=tk.W, pady=2)

# Min payment due on
lday_label_t3 = tk.Label(bal_frame_t3, text="due on")
lday_label_t3.grid(row=0, column=2, sticky=tk.W, pady=2)
lday_entry_t3 = DateEntry(bal_frame_t3, width=12, background="white", foreground="black")
lday_entry_t3.grid(row=0, column=3, sticky=tk.W, pady=2)

# Current balance

currbal_label = tk.Label(bal_frame_t3, text= "Current balance:")
currbal_label.grid(row=1, column=0, sticky=tk.W, pady=2)
currbal_entry_t3 = tk.Entry(bal_frame_t3, width=10)
currbal_entry_t3.focus_set()
currbal_entry_t3.grid(row=1, column=1, sticky=tk.W, pady=2)

# Day
day_label_t3 = tk.Label(bal_frame_t3, text= "Day due:")
day_label_t3.grid(row=2, column=0, sticky=tk.W, pady=2)
day_entry_t3 = tk.Entry(bal_frame_t3, width=10)
day_entry_t3.focus_set()
day_entry_t3.grid(row=2, column=1, sticky=tk.W, pady=2)

# End of month text
eom_text_t3 = "(Leave blank if payment is due before the end of the month)"
eom_label_t3 = tk.Label(bal_frame_t3, text=eom_text_t3)
eom_label_t3.place(x=100, y=250)

def open_file_dialog_t3():
    file_path = filedialog.askopenfilename()
    if file_path:
        file_label_t3.config(text="Selected File: " + file_path)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        # Remove brackets and commas
        data = [item for sublist in data for item in sublist]
        insertOutlineData_t3(data)

    else:
        file_label_t3.config(text="No file selected")

# Upload Outline Frame
outline_frame_t3 = tk.LabelFrame(tab3, text="Upload Outline", padx=2, pady=2)
outline_frame_t3.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Upload Button
upload_button_t3 = tk.Button(outline_frame_t3, text="Upload", command=open_file_dialog_t3)
upload_button_t3.grid(row=0, column=0, padx=5, pady=5)

# Label to display selected file name
file_label_t3 = tk.Label(outline_frame_t3, text="No file selected")
file_label_t3.grid(row=0, column=1, padx=5, pady=5)


def createVPA(firm, name, addr, csz, email, client, court, case, minpay, next_pay, day, bal):
    if firm == 'NWTLG' and email != "":
        doc_path = "F:\\DocumentCreator\\VPA\\nwvpa.docx"
    elif firm == 'NWTLG' and email == "":
        doc_path = "F:\\DocumentCreator\\VPA\\nwvpa_no_email.docx"
    elif firm == 'HAD' and email != "":
        doc_path = "F:\\DocumentCreator\\VPA\\hadvpa.docx"
    else:
        doc_path = "F:\\DocumentCreator\\VPA\\hadvpa_no_email.docx"

    doc = DocxTemplate(doc_path)

    if day == "":
        day = 'end'
    else:
        day = day + suffix(day)

    email = '*Sent via email: ' + email

    context = {'name': name, 'addr': addr, 'csz': csz, 'email': email, 'client': client, 'court': court, 'case': case,
               'min': formatNumberWithCommas(float(minpay)), 'min_day': day,  'next_pay': convertDate_t2(next_pay), 'bal': formatNumberWithCommas(float(bal)), 'init': initials_input.get()}
    
    doc.render(context)
    fileName = genFile(name, 'VPA Letter')
    dir_path = "C:\\Users\\mark\\Automations\\cure\\"
    file_path = dir_path + fileName
    print(file_path)
    doc.save(file_path)

def submitFunction_t3():
    firm = firm_var_t3.get()
    name = name_entry_t3.get()
    addr = addr_entry_t3.get()
    csz = csz_entry_t3.get()
    email = email_entry_t3.get()
    client = client_entry_t3.get()
    court = court_entry_t3.get()
    case = case_entry_t3.get()
    minpay = min_entry_t3.get()
    next_pay = lday_entry_t3.get()
    day = day_entry_t3.get()
    bal = currbal_entry_t3.get()
    
    createVPA(firm, name, addr, csz, email, client, court, case, minpay, next_pay, day, bal)

def clearFunction_t3():
    name_entry_t3.delete(0, 'end')
    addr_entry_t3.delete(0, 'end')
    csz_entry_t3.delete(0, 'end')
    email_entry_t3.delete(0, 'end')
    client_entry_t3.delete(0, 'end')
    court_entry_t3.delete(0, 'end')
    case_entry_t3.delete(0, 'end')
    min_entry_t3.delete(0, 'end')
    lday_entry_t3.delete(0, 'end')
    day_entry_t3.delete(0, 'end')
    currbal_entry_t3.delete(0, 'end')

def insertOutlineData_t3(outlineData):
    if outlineData[0] == 'NWTLG':
        firm_var_t3.set("NWTLG")
    elif outlineData[0] == 'HAD':
        firm_var_t3.set("HAD")

    name_entry_t3.delete(0, 'end')
    name_entry_t3.insert(0, outlineData[1])
    addr_entry_t3.delete(0, 'end')
    addr_entry_t3.insert(0, outlineData[2])
    csz_entry_t3.delete(0, 'end')
    csz_entry_t3.insert(0, outlineData[3] + ', '+ outlineData[4] + ' ' + outlineData[5])
    email_entry_t3.delete(0, 'end')
    email_entry_t3.insert(0, outlineData[6])
    client_entry_t3.delete(0, 'end')
    client_entry_t3.insert(0, modify_string(outlineData[7]))
    court_entry_t3.delete(0, 'end')
    court_entry_t3.insert(0, outlineData[8])
    case_entry_t3.delete(0, 'end')
    case_entry_t3.insert(0, outlineData[9])
    min_entry_t3.delete(0, 'end')
    min_entry_t3.insert(0, clean_number(outlineData[11]))
    day_entry_t3.delete(0, 'end')
    day_entry_t3.insert(0, outlineData[12])


# Create a frame for the buttons
button_frame_t3 = tk.Frame(tab3)
button_frame_t3.pack(pady=10)

# Create the Clear button
clear_button_t3 = tk.Button(button_frame_t3, text="Clear", command=clearFunction_t3)
clear_button_t3.pack(side=tk.LEFT, padx=(0, 5))

# Create the Submit button
submit_button_t3 = tk.Button(button_frame_t3, text="Save", command=submitFunction_t3)
submit_button_t3.pack(side=tk.LEFT, padx=(5, 0))

###################################
#
# PAYOFF TAB
#
##################################
# Firm section
firm_frame = tk.LabelFrame(tab4, text="Firm", padx=2, pady=2)
firm_frame.pack(padx=3, pady=3, fill="both")

firm_var_t4 = tk.StringVar()

had_radio_t4 = tk.Radiobutton(firm_frame, text="HAD", variable=firm_var_t4, value="HAD", tristatevalue=0)
had_radio_t4.pack(side=tk.LEFT, padx=5)

nwltg_radio_t4 = tk.Radiobutton(firm_frame, text="NWTLG", variable=firm_var_t4, value="NWTLG", tristatevalue=0)
nwltg_radio_t4.pack(side=tk.LEFT, padx=5)

# Information Section
info_frame_t4 = tk.LabelFrame(tab4, text="Information", padx=2, pady=2)
info_frame_t4.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Name field
name_label_t4 = tk.Label(info_frame_t4, text="Agent Name:")
name_label_t4.grid(row=0, column=0, sticky=tk.W, pady=2)
name_entry_t4 = tk.Entry(info_frame_t4, width=30)
name_entry_t4.focus_set()
name_entry_t4.grid(row=0, column=1, sticky=tk.W, pady=2)

# Job Title field
title_label_t4 = tk.Label(info_frame_t4, text="Job Title:")
title_label_t4.grid(row=1, column=0, sticky=tk.W, pady=2)
title_entry_t4 = tk.Entry(info_frame_t4, width=30)
title_entry_t4.focus_set()
title_entry_t4.grid(row=1, column=1, sticky=tk.W, pady=2)

# Title company field
comp_label_t4 = tk.Label(info_frame_t4, text="Title Company:")
comp_label_t4.grid(row=2, column=0, sticky=tk.W, pady=2)
comp_entry_t4 = tk.Entry(info_frame_t4, width=30)
comp_entry_t4.focus_set()
comp_entry_t4.grid(row=2, column=1, sticky=tk.W, pady=2)

# Phone field
phone_label_t4 = tk.Label(info_frame_t4, text="Phone:")
phone_label_t4.grid(row=3, column=0, sticky=tk.W, pady=2)
phone_entry_t4 = tk.Entry(info_frame_t4, width=30)
phone_entry_t4.focus_set()
phone_entry_t4.grid(row=3, column=1, sticky=tk.W, pady=2)

# Phone field
fax_label_t4 = tk.Label(info_frame_t4, text="Fax:")
fax_label_t4.grid(row=4, column=0, sticky=tk.W, pady=2)
fax_entry_t4 = tk.Entry(info_frame_t4, width=30)
fax_entry_t4.focus_set()
fax_entry_t4.grid(row=4, column=1, sticky=tk.W, pady=2)

# Email field
email_label_t4 = tk.Label(info_frame_t4, text="Email:")
email_label_t4.grid(row=5, column=0, sticky=tk.W, pady=2)
email_entry_t4 = tk.Entry(info_frame_t4, width=30)
email_entry_t4.focus_set()
email_entry_t4.grid(row=5, column=1, sticky=tk.W, pady=2)

# Plaintiff field
ptiff_label_t4 = tk.Label(info_frame_t4, text="Plaintiff:")
ptiff_label_t4.grid(row=0, column=2, sticky=tk.W, pady=2)
ptiff_entry_t4 = tk.Entry(info_frame_t4, width=30)
ptiff_entry_t4.focus_set()
ptiff_entry_t4.grid(row=0, column=3, sticky=tk.W, pady=2)

# Defendant field
deff_label_t4 = tk.Label(info_frame_t4, text="Defendant:")
deff_label_t4.grid(row=1, column=2, sticky=tk.W, pady=2)
deff_entry_t4 = tk.Entry(info_frame_t4, width=30)
deff_entry_t4.focus_set()
deff_entry_t4.grid(row=1, column=3, sticky=tk.W, pady=2)

# Court field
court_label_t4 = tk.Label(info_frame_t4, text="Court:")
court_label_t4.grid(row=2, column=2, sticky=tk.W, pady=2)
court_entry_t4 = tk.Entry(info_frame_t4, width=30)
court_entry_t4.focus_set()
court_entry_t4.grid(row=2, column=3, sticky=tk.W, pady=2)

# Case number field
case_label_t4 = tk.Label(info_frame_t4, text="Case No:")
case_label_t4.grid(row=3, column=2, sticky=tk.W, pady=2)
case_entry_t4 = tk.Entry(info_frame_t4, width=20)
case_entry_t4.focus_set()
case_entry_t4.grid(row=3, column=3, sticky=tk.W, pady=2)

# State label
state_label_t4 = tk.Label(info_frame_t4, text="State:")
state_label_t4.grid(row=4, column=2, sticky=tk.W, pady=2)

# File number field
file_label_t4 = tk.Label(info_frame_t4, text="File No:")
file_label_t4.grid(row=4, column=2, sticky=tk.W, pady=2)
file_entry_t4 = tk.Entry(info_frame_t4, width=20)
file_entry_t4.focus_set()
file_entry_t4.grid(row=4, column=3, sticky=tk.W, pady=2)

# State radio buttons
state_var_t4 = tk.StringVar()
state_var_t4.set("OR")  # Set default value to OR

radio_frame_t4 = tk.Frame(info_frame_t4)
radio_frame_t4.grid(row=5, column=3, sticky=tk.W, pady=2)

or_radio_t4 = tk.Radiobutton(radio_frame_t4, text="OR", variable=state_var_t4, value="OR")
or_radio_t4.pack(side=tk.LEFT)

wa_radio_t4 = tk.Radiobutton(radio_frame_t4, text="WA", variable=state_var_t4, value="WA")
wa_radio_t4.pack(side=tk.LEFT)

# Payoff Frame
po_frame = tk.LabelFrame(tab4, text="Remaining balance", padx=2, pady=2)
po_frame.pack(side=tk.TOP, padx=3, pady=3, fill="both")

#Current balance
curbal_label = tk.Label(po_frame, text="Current balance:")
curbal_label.grid(row=0, column=0, sticky=tk.W, pady=2)
curbal_entry = tk.Entry(po_frame, width=10)
curbal_entry.focus_set()
curbal_entry.grid(row=0, column=1, sticky=tk.W, pady=2)

#Balance owing
balowe_label = tk.Label(po_frame, text="Balance owing:")
balowe_label.grid(row=1, column=0, sticky=tk.W, pady=2)
balowe_entry = tk.Entry(po_frame, width=10)
balowe_entry.focus_set()
balowe_entry.grid(row=1, column=1, sticky=tk.W, pady=2)

#Good through
good_label = tk.Label(po_frame, text="Good through:")
good_label.grid(row=1, column=2, sticky=tk.W, pady=2)
good_entry = DateEntry(po_frame, width=12, background="white", foreground="black")
good_entry.grid(row=1, column=3, sticky=tk.W, pady=2)

#Per diem
diem_label = tk.Label(po_frame, text="Per diem:")
diem_label.grid(row=2, column=0, sticky=tk.W, pady=2)
diem_entry = tk.Entry(po_frame, width=10)
diem_entry.focus_set()
diem_entry.grid(row=2, column=1, sticky=tk.W, pady=2)

#Judgment
jmt_label = tk.Label(po_frame, text="Judgment Entry:")
jmt_label.grid(row=2, column=2, sticky=tk.W, pady=2)
jmt_entry = DateEntry(po_frame, width=12, background="white", foreground="black")
jmt_entry.grid(row=2, column=3, sticky=tk.W, pady=2)

def open_file_dialog_t4():
    file_path = filedialog.askopenfilename()
    if file_path:
        file_label_t4.config(text="Selected File: " + file_path)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        # Remove brackets and commas
        data = [item for sublist in data for item in sublist]
        insertOutlineData_t4(data)

    else:
        file_label_t4.config(text="No file selected")

# Upload Outline Frame
outline_frame_t4 = tk.LabelFrame(tab4, text="Upload Outline", padx=2, pady=2)
outline_frame_t4.pack(side=tk.TOP, padx=3, pady=3, fill="both")

# Upload Button
upload_button_t4 = tk.Button(outline_frame_t4, text="Upload", command=open_file_dialog_t4)
upload_button_t4.grid(row=0, column=0, padx=5, pady=5)

# Label to display selected file name
file_label_t4 = tk.Label(outline_frame_t4, text="No file selected")
file_label_t4.grid(row=0, column=1, padx=5, pady=5)


# Create a frame for the buttons
button_frame_t4 = tk.Frame(tab4)
button_frame_t4.pack(pady=10)

def add10(date_string):
    # Convert the input date string to a datetime object
    date_format = "%m/%d/%y"
    date_obj = datetime.strptime(date_string, date_format)

    # Add 10 days to the datetime object
    new_date = date_obj + timedelta(days=10)

    # Convert the new date back to a string in the original format
    new_date_string = new_date.strftime(date_format)
    return new_date_string

def fixDate(input_date):
    try:
        # Parse the input date string using the datetime.strptime function
        date_obj = datetime.strptime(input_date, "%m/%d/%y")

        # Format the date object in the desired output format
        output_date = date_obj.strftime("%B %d, %Y").replace(" 0", " ")

        return output_date
    except ValueError:
        return "Invalid date format. Please use 'mm/dd/yy'."

def createPayoff(firm, state, name, title, co, phone, fax, email, ptiff, deff, court, case, curbal, balowe, gdate, pd, jdate, fileno):
    if pd == "" or pd == "0":
        pdlong = 'This offer expires on ' + fixDate(add10(gdate)) + ' . If the closing cannot be accomplished by this date, the payoff is no longer valid and we will resume collection efforts.'
        pd = '0.00'
    else:
        pdlong = "This offer expires on " + fixDate(add10(gdate)) + ". If the above date cannot be met, the per diem is $"+ formatNumberWithCommas(pd) + " for up to ten days after your expected closing. If the closing cannot be accomplished by the later date, we may resume collection efforts."

    if state == 'OR':
        atty = 'David DeBlasio'
    else:
        atty = "Ann L Fisher"

    if firm == 'NWTLG':
        doc_path = "F:\\DocumentCreator\\Payoff\\nwpo.docx"
    else:
        doc_path = "F:\\DocumentCreator\\Payoff\\hadpo.docx"

    doc = DocxTemplate(doc_path)

    context = {'name': name, 'title': title, 'company': co, 'phone': phone, 'fax': fax, 'email': email,
              'client': ptiff, 'def': deff, 'court': court, 'case': case, 'curbal': formatNumberWithCommas(str(curbal)), 'pd': pdlong, 'pdiem': pd, 'bal': formatNumberWithCommas(str(balowe)),
             'date': fixDate(gdate), 'jdate': fixDate(jdate), 'Atty': atty, 'init': initials_input.get(), 'file': fileno}
    
    doc.render(context)
    fileName = genFile(deff, 'Payoff Letter')
    dir_path = "C:\\Users\\mark\\Automations\\cure\\"
    file_path = dir_path + fileName
    print(file_path)
    doc.save(file_path)

def submitFunction_t4():
    firm = firm_var_t4.get()
    state = state_var_t4.get()
    name = name_entry_t4.get()
    title = title_entry_t4.get()
    co = comp_entry_t4.get()
    phone = phone_entry_t4.get()
    fax = fax_entry_t4.get()
    email = email_entry_t4.get()
    ptiff = ptiff_entry_t4.get()
    deff = deff_entry_t4.get()
    court = court_entry_t4.get()
    case = case_entry_t4.get()
    curbal = curbal_entry.get()
    balowe = balowe_entry.get()
    gdate = good_entry.get()
    pd = diem_entry.get()
    jdate = jmt_entry.get()
    fileno = file_entry_t4.get()

    createPayoff(firm, state, name, title, co, phone, fax, email, ptiff, deff, court, case, curbal, balowe, gdate, pd, jdate, fileno)

def removeSPGLF(input_string):
    if input_string.startswith('SPGLF-'):
        return input_string[len('SPGLF-'):]
    else:
        return input_string

def insertOutlineData_t4(outlineData):
    if outlineData[0] == 'NWTLG':
        firm_var_t4.set("NWTLG")
    elif outlineData[0] == 'HAD':
        firm_var_t4.set("HAD")

    ptiff_entry_t4.delete(0, 'end')
    ptiff_entry_t4.insert(0, modify_string(outlineData[7]))
    deff_entry_t4.delete(0, 'end')
    deff_entry_t4.insert(0, outlineData[1])
    court_entry_t4.delete(0, 'end')
    court_entry_t4.insert(0, outlineData[8])
    case_entry_t4.delete(0, 'end')
    case_entry_t4.insert(0, outlineData[9])

    if outlineData[13] == 'OR':
        state_var_t4.set("OR")
    else:
        state_var_t4.set("WA")

    file_entry_t4.delete(0, 'end')
    file_entry_t4.insert(0, removeSPGLF(outlineData[14]))

    def fixJmtDate(input_date):
        parts = input_date.split('/')
        if len(parts) == 3:
            year = parts[2][2:]
            return '/'.join([parts[0], parts[1], year])
        else:
            print(input_date)
            return input_date
        
    jmt_entry.delete(0, 'end') 
    jmt_entry.insert(0, fixJmtDate(outlineData[15]))

def clearFunction_t4():
    name_entry_t4.delete(0, 'end')
    title_entry_t4.delete(0, 'end')
    comp_entry_t4.delete(0, 'end')
    phone_entry_t4.delete(0, 'end')
    fax_entry_t4.delete(0, 'end')
    email_entry_t4.delete(0, 'end')
    ptiff_entry_t4.delete(0, 'end')
    deff_entry_t4.delete(0, 'end')
    court_entry_t4.delete(0, 'end')
    case_entry_t4.delete(0, 'end')
    curbal_entry.delete(0, 'end')
    balowe_entry.delete(0, 'end')
    good_entry.delete(0, 'end')
    diem_entry.delete(0, 'end')
    jmt_entry.delete(0, 'end')
    file_entry_t4.delete(0, 'end')

# Create the Clear button
clear_button_t4 = tk.Button(button_frame_t4, text="Clear", command=clearFunction_t4)
clear_button_t4.pack(side=tk.LEFT, padx=(0, 5))

# Create the Submit button
submit_button_t4 = tk.Button(button_frame_t4, text="Print", command=submitFunction_t4)
submit_button_t4.pack(side=tk.LEFT, padx=(5, 0))

#######################
#
#   CHECKS TAB
#
########################

def choose_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    file_name_label.config(text="File Name: " + file_path)

def convertArray(array):
    print('stringarr')
    print(array)
    string_array = array.astype(str)
    return string_array

def convertDate(date_string):
    date_string = date_string.replace('/', '.')  # Replace slashes with dots
    parts = date_string.split('.')  # Split the string into parts
    
    # Add leading zeros to each part
    formatted_parts = [part.zfill(2) for part in parts]
    
    # Join the parts with dots to form the final formatted date
    formatted_date = '.'.join(formatted_parts)
    
    return formatted_date

def addZero(string_array):
    for i in range(len(string_array)):
        if len(string_array[i]) == 7:
            string_array[i] = '0' + string_array[i]
    return string_array

def recalibrate():
    chrome_process = subprocess.Popen(['C:\Program Files (x86)\Google\Chrome\Application\chrome.exe', '--new-window', '--start-maximized', '--force-desktop', 'https://app.clio.com/nc/#/'])
    time.sleep(1)
    chrome_window = win32gui.GetForegroundWindow()
    win32gui.SetForegroundWindow(chrome_window)
    pyautogui.hotkey('win', 'up')  # Maximize the Chrome window
    # Initialize a list to store mouse click data
    click_data = []

    # Define the callback function for mouse clicks
    def on_click(x, y, button, pressed):
        if pressed:
            click_data.append((x, y))

        # If we have collected 5 data points, stop listening
        if len(click_data) == 5:
            chrome_process.kill()
            return False

    # Set up the mouse listener
    with Listener(on_click=on_click) as listener:
        listener.join()

   # Create a DataFrame from the click data
    df = pd.DataFrame(click_data, columns=['X', 'Y'])

    # Write the DataFrame to the Excel file
    df.to_excel('calibration.xlsx', sheet_name='Sheet1', index=False)


def submit():
    file_path = file_name_label.cget("text").split(": ")[1]
    df = pd.read_excel(file_path)
    filtered_df = df.loc[df.iloc[:, 0].notnull() & df.iloc[:, 3].notnull()]
    filtered_df = filtered_df.iloc[:, [0, 3]]

    fileNumber = filtered_df.iloc[:, 0].astype(str).values
    amountPayed = filtered_df.iloc[:, 1].astype(str).values

    print(fileNumber)
    # Get input fields
    type = radio_var.get()
    date = date_var.get()
    date = convertDate(date)
    toClio(fileNumber, amountPayed, type, date)

def toClio(fileNumber, amountPayed, type, date):
    df = pd.read_excel('calibration.xlsx')
    keyboard  = Controller()
    subprocess.Popen(['C:\Program Files (x86)\Google\Chrome\Application\chrome.exe', '--new-window', '--start-maximized', '--force-desktop', 'https://app.clio.com/nc/#/'])
    time.sleep(2)
    chrome_window = win32gui.GetForegroundWindow()
    win32gui.SetForegroundWindow(chrome_window)
    pyautogui.hotkey('win', 'up')  # Maximize the Chrome window
    time.sleep(5)
    for i in range(len(fileNumber)):
        pyautogui.click(x=df.iloc[0, 0], y=df.iloc[0, 1])
        print(df.iloc[0, 1])
        print(fileNumber[i])
        pyautogui.typewrite(fileNumber[i])
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        time.sleep(2)

        for _ in range(26):
            keyboard.press(Key.tab)
            keyboard.release(Key.tab)
        time.sleep(2)
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        time.sleep(3)
        pyautogui.click(x=df.iloc[2, 0], y=df.iloc[2, 1])
        time.sleep(2)
        pyautogui.click(x=df.iloc[3, 0], y=df.iloc[3, 1])
        time.sleep(2)
        pyautogui.click(x=df.iloc[4, 0], y=df.iloc[4, 1])
        pyautogui.typewrite(amountPayed[i])
        for _ in range(6):
            keyboard.press(Key.tab)
            keyboard.release(Key.tab)
        # Directpay
        if type == 'directPay':
            src = 'CLIENT - DIRECT PAYMENT'
            ref = 'SPGLF ' + date + ' CONTINGENCY'
            pyautogui.typewrite(src)
            for _ in range(4):
                keyboard.press(Key.tab)
                keyboard.release(Key.tab)
            pyautogui.typewrite(ref)
            for _ in range(2):
                keyboard.press(Key.tab)
                keyboard.release(Key.tab)
            keyboard.press(Key.enter)
            keyboard.release(Key.enter)
            time.sleep(2)
        # Supplemental
        elif type == 'supplemental':
            src = 'FIRM RCPT - ' + date + ' SUPPLEMENTAL REMIT'
            pyautogui.typewrite(src)
            for _ in range(6):
                keyboard.press(Key.tab)
                keyboard.release(Key.tab) 
            keyboard.press(Key.enter)
            keyboard.release(Key.enter)
            time.sleep(2)

        # firm RCPT
        elif type == 'firmRCPT':
            src = 'FIRM RCPT - ' + date + ' REMIT'
            pyautogui.typewrite(src)
            for _ in range(6):
                keyboard.press(Key.tab)
                keyboard.release(Key.tab) 
            keyboard.press(Key.enter)
            keyboard.release(Key.enter)
            time.sleep(2)

# SPREADSHEET SELECTION FRAME
checkFrame = tk.LabelFrame(tab5, text="Spreadsheet Selection", padx=2, pady=2)
checkFrame.pack(padx=3, pady=1, fill="both")

# Choose File Button
choose_file_button = tk.Button(checkFrame, text="Choose File", command=choose_file)
choose_file_button.grid(row=0, column=0, pady=20, sticky='w')

# File Name Label
file_name_label = tk.Label(checkFrame, text="File Name:")
file_name_label.grid(row=0, column=1, sticky='w')

# TRANSACTION TYPE FRAME
transacFrame = tk.LabelFrame(tab5, text="Transaction Type", padx=2, pady=2)
transacFrame.pack(padx=3, pady=1, fill="both")

# Radio Buttons
radio_var = tk.StringVar()
radio_var.set("none")

radio1 = tk.Radiobutton(transacFrame, text="Direct Pay", variable=radio_var, value="directPay")
radio1.grid(row=1, column=0, sticky='w')

radio2 = tk.Radiobutton(transacFrame, text="Firm Receipt", variable=radio_var, value="firmRCPT")
radio2.grid(row=2, column=0, sticky='w')

radio3 = tk.Radiobutton(transacFrame, text="Supplemental", variable=radio_var, value="supplemental")
radio3.grid(row=3, column=0, sticky='w')

radio4 = tk.Radiobutton(transacFrame, text="Other", variable=radio_var, value="other")
radio4.grid(row=4, column=0, sticky='w')

other_entry = tk.Entry(transacFrame)
other_entry.grid(row=4, column=1, sticky='w')

# Date
dateLabel = tk.Label(transacFrame, text= 'Remit date:')
dateLabel.grid(row=5, column = 0, stick='w')
date_var = DateEntry(transacFrame, width=12, background="white", foreground="black")
date_var.grid(row=5, column = 1, stick='w')

def handle_radio():
    if radio_var.get() == "other":
        other_entry.config(state="normal")
    else:
        other_entry.delete(0, tk.END)
        other_entry.config(state="disabled")

radio_var.trace("w", lambda *args: handle_radio())

# Create a frame for the buttons
button_frame_t5 = tk.Frame(tab5)
button_frame_t5.pack(pady=10)

# Submit Button
submit_button = tk.Button(button_frame_t5, text="Enter Transactions", command=submit)
submit_button.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

# Calibrate button
calibrate_button = tk.Button(button_frame_t5, text="Calibrate Mouse", command=recalibrate)
calibrate_button.pack(side=tk.LEFT, anchor="w", padx=(5, 0))

window.mainloop()
